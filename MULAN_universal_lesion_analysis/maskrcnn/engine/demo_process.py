# Ke Yan, Imaging Biomarkers and Computer-Aided Diagnosis Laboratory,
# National Institutes of Health Clinical Center, July 2019
"""Procedure in the demo mode"""
import os
import numpy as np
from time import time
import pickle
import torch
import nibabel as nib
import nrrd
from tqdm import tqdm
import cv2
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
# from scipy import interpolate

from maskrcnn.config import cfg
from maskrcnn.data.datasets.load_ct_img import load_prep_img
from maskrcnn.structures.image_list import to_image_list
from maskrcnn.data.datasets.evaluation.DeepLesion.post_process import post_process_results
from maskrcnn.data.datasets.load_ct_img import windowing, windowing_rev
from maskrcnn.utils.draw import draw_results


def exec_model(model):
    """test model on user-provided data, instead of the preset DeepLesion dataset"""
    if cfg.MODEL.TAG_ON:
        import_tag_data()
    model.eval()
    device = torch.device(cfg.MODEL.DEVICE)

    while True:
        info = "Please input the path of a nifti or nrrd CT volume >> "
        while True:
            path = input(info)
            if not os.path.exists(path):
                print('file does not exist!')
                continue
            # try:
            print(path)
            if path.split(".")[-1] == 'nrrd':
                print('reading nrrd file ...')
                data_is_nrrd = True
                nrrd_data = nrrd.read(path)
                print(len(nrrd_data))
            else:
                print('reading nifti file ...')
                nifti_data = nib.load(path)
            break
            # except:
            #     print('load nifti file error!')

        while True:
            win_sel = input('Window to show, 1:soft tissue, 2:lung, 3: bone >> ')
            if win_sel not in ['1', '2', '3']:
                continue
            win_show = [[-175, 275], [-1500, 500], [-500, 1300]]
            win_show = win_show[int(win_sel)-1]
            break

        if data_is_nrrd:
            vol, spacing, slice_intv = load_preprocess_nrrd(nrrd_data)
        else:
            vol, spacing, slice_intv = load_preprocess_nifti(nifti_data)
            
        slice_num_per_run = max(1, int(float(cfg.TEST.TEST_SLICE_INTV_MM)/slice_intv+.5))
        num_total_slice = vol.shape[2]

        total_time = 0
        output_dir = os.path.join(cfg.RESULTS_DIR,path.replace(os.sep, '_'))
        if not os.path.exists(output_dir):
            os.mkdir(output_dir)

        # from NRRD files - get patient id (P__)
        image_fn = path.split("/")[-1]
        pid = image_fn.split("_")[0]

        # number of slices for each filename is saved in under the pickle file below
        # load number of slides for each file name into slices_to_process
        dbfile = open('/home/tester/jessica/MULAN_BC/MULAN_universal_lesion_analysis/maskrcnn/data/datasets/lesion_slice_dict.pkl', 'rb')     
        lesion_slice_dict = pickle.load(dbfile)
        # get the full range of CT slices to pass thru the model 
        slices_to_process = lesion_slice_dict[image_fn]

        # load and preprocess mask data (segmentation mask gt)
        mask_path = "/".join(path.split('/')[:-1]) + "/" + str(pid) + "_mask.nrrd"
        mask_data = nrrd.read(mask_path)
        mask, spacing, slice_intv = load_preprocess_nrrd(mask_data, False)
        mask = np.transpose(mask, (2, 0, 1))

        msgs_all = []
        print('predicting ...')
        for slice_idx in tqdm(slices_to_process):
            ims, im_np, im_scale, crop = get_ims(slice_idx, vol, spacing, slice_intv)
            im_list = to_image_list(ims, cfg.DATALOADER.SIZE_DIVISIBILITY).to(device)
            start_time = time()
            with torch.no_grad():
                result = model(im_list)
            result = [o.to("cpu") for o in result]

            info = {'spacing': spacing, 'im_scale': im_scale}
            post_process_results(result[0], info)
            total_time += time() - start_time
            overlay, msgs = gen_output(im_np, result[0], info, win_show)
            
            # NOTE: This section is modified for the Breast CT scans (NRRD files) pipeline
            # section: vizualisation pipeline to draw predicted contours + gt contour + gt box
            # predicted boxes are drawn using the original code in the gen_output function
            np_lesion = np.argwhere(mask[slice_idx]==1)
            if len(np_lesion) == 0:
                continue
            im_scale = (overlay.shape[1])/(512*2)
            spacing = im_scale*2

            # 1. drawing predicted segmentations
            for box in result:
                pred_masks = box.get_field("mask")
                for pred_mask in pred_masks:
                    interpolated_mask = torch.nn.functional.interpolate(pred_mask.unsqueeze(0).unsqueeze(0), scale_factor=spacing)
                    interpolated_mask = interpolated_mask[0][0].cpu().numpy()
                    interpolated_mask = np.where(interpolated_mask == 1, 255, interpolated_mask).astype(np.uint8)
                    idx = cv2.findContours(interpolated_mask,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)[0]
                    idx = np.array(idx)
                    
                    # if contour can be found - draw contour
                    if idx.ndim == 4: 
                        idx = np.reshape(idx, (idx.shape[1], 2))
                        out = np.zeros_like(interpolated_mask)
                        out[idx[:,1],idx[:,0]] = 255
                        interpolated_mask = out
                        interpolated_mask = np.pad(interpolated_mask, pad_width=((0, overlay.shape[0]-overlay.shape[1]), (0, 0)))
                        interpolated_mask = np.ma.make_mask(interpolated_mask)
                        overlay[interpolated_mask] = np.array([0,0,255])

            # 2a. getting GT mask 
            curr_mask = mask[slice_idx].copy()
            torch_mask = torch.from_numpy(curr_mask.astype("float32"))
            interpolated_mask = torch.nn.functional.interpolate(torch_mask.unsqueeze(0).unsqueeze(0), scale_factor=spacing)
            interpolated_mask = interpolated_mask[0][0].cpu().numpy()

            # 2b. getting contour of GT mask
            interpolated_mask = np.where(interpolated_mask == 1, 255, interpolated_mask).astype(np.uint8)
            idx = cv2.findContours(interpolated_mask,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)[0]
            idx = np.array(idx)
            # if contour is found - draw contour
            if idx.ndim == 4: 
                idx = np.reshape(idx, (idx.shape[1], 2))
                out = np.zeros_like(interpolated_mask)
                out[idx[:,1],idx[:,0]] = 255
                interpolated_mask = out
            # NOTE: else - interpolated_mask stays as the mask and the entire mask will be shaded in the final vizualisation
        
            interpolated_mask = np.pad(interpolated_mask, pad_width=((0, overlay.shape[0]-overlay.shape[1]), (0, 0)))
            interpolated_mask = np.ma.make_mask(interpolated_mask)
            overlay[interpolated_mask] = np.array([255,0,0])

            ## 2c. draw gt box - changed to drawing contour instead
            # left = round((min(np_lesion[:, 1])-1)*spacing)
            # down = round((min(np_lesion[:, 0])-1)*spacing)
            # right = round((max(np_lesion[:, 1])-1)*spacing)
            # up = round((max(np_lesion[:, 0])-1)*spacing)
            #overlay = cv2.rectangle(overlay, (left, down), (right, up), (255, 0, 0), 1)
        

            # 3. configure directory to save the vizualizations
            viz_folder = "/home/tester/jessica/MULAN_BC/MULAN_universal_lesion_analysis/viz/" + str(cfg.SAVED_WEIGHTS.split(" ")[-1]) + "/" + str(pid)
            os.makedirs(viz_folder, exist_ok=True)
            viz_fn = viz_folder + "/" +  str(slice_idx) + ".png"

            # 3b. plot and save visualizations
            fig, ax = plt.subplots()
            ax.imshow(overlay)
            ax.figure.savefig(viz_fn, dpi=300)
            #cv2.imwrite(output_fn, overlay)

            ## NOTE: End of modified section

            msgs_all.append('slice %d\r\n' % (slice_idx+1))
            for msg in msgs:
                msgs_all.append(msg+'\r\n')
            msgs_all.append('\r\n')

        with open(os.path.join(output_dir, 'results.txt'), 'w') as f:
            f.writelines(msgs_all)

        print('result images and text saved to', output_dir)
        print('processing time: %d ms per slice' % int(1000.*total_time/len(slices_to_process)))


def import_tag_data():
    cellname = lambda row, col: '%s%d' % (chr(ord('A') + col - 1), row)
    fn = os.path.join(cfg.PROGDAT_DIR, '%s_%s.xlsx' % ('test_handlabeled', cfg.EXP_NAME))

    wb = load_workbook(fn)
    sheet = wb.get_active_sheet()
    tags = []
    thresolds = []
    for p in range(2, sheet.max_row):
        tags.append(sheet[cellname(p, 1)].value)
        thresolds.append(float(sheet[cellname(p, 8)].value))
    assert tags == cfg.runtime_info.tag_list
    cfg.runtime_info.tag_sel_val = torch.tensor(thresolds).to(torch.float)


def load_preprocess_nifti(data):
    vol = (data.get_fdata().astype('int32') + 32768).astype('uint16')  # to be consistent with png files
    # spacing = -data.get_affine()[0,1]
    # slice_intv = -data.get_affine()[2,2]
    aff = data.affine[:3, :3]
    spacing = np.abs(aff[:2, :2]).max()
    slice_intv = np.abs(aff[2, 2])

    # TODO: Ad-hoc code for normalizing the orientation of the volume.
    # The aim is to make vol[:,:,i] an supine right-left slice
    # It works for the authors' data, but maybe not suitable for some kinds of nifti files
    if np.abs(aff[0, 0]) > np.abs(aff[0, 1]):
        vol = np.transpose(vol, (1, 0, 2))
        aff = aff[[1, 0, 2], :]
    if np.max(aff[0, :2]) > 0:
        vol = vol[::-1, :, :]
    if np.max(aff[1, :2]) > 0:
        vol = vol[:, ::-1, :]
    return vol, spacing, slice_intv

def load_preprocess_nrrd(data, image_conversion=True):
    if image_conversion:
        vol = (data[0].astype('int32') + 32768).astype('uint16')  # to be consistent with png files
    else:
        vol = data[0]
    # spacing = -data.get_affine()[0,1]
    # slice_intv = -data.get_affine()[2,2]
    aff = data[1]['space directions'][:3, :3]
    spacing = np.abs(aff[:2, :2]).max()
    slice_intv = np.abs(aff[2, 2])

    if np.abs(aff[0, 0]) > np.abs(aff[0, 1]):
        vol = np.transpose(vol, (1, 0, 2))
        aff = aff[[1, 0, 2], :]
    if np.max(aff[0, :2]) > 0:
        vol = vol[::-1, :, :]
    if np.max(aff[1, :2]) > 0:
        vol = vol[:, ::-1, :]

    return vol, spacing, slice_intv



def get_ims(slice_idx, vol, spacing, slice_intv):
    num_slice = cfg.INPUT.NUM_SLICES * cfg.INPUT.NUM_IMAGES_3DCE # == 9 following the settings in default.py
    im_np, im_scale, crop = load_prep_img(vol, slice_idx, spacing, slice_intv,
                                          cfg.INPUT.IMG_DO_CLIP, num_slice=num_slice)
    im = im_np - cfg.INPUT.PIXEL_MEAN
    im = torch.from_numpy(im.transpose((2, 0, 1))).to(dtype=torch.float)
    ims = im.split(cfg.INPUT.NUM_IMAGES_3DCE)
    return ims, im_np[:, :, int(num_slice/2)+1], im_scale, crop


def gen_output(im, result, info, win_show):
    im = windowing_rev(im, cfg.INPUT.WINDOWING)
    im = windowing(im, win_show).astype('uint8')
    im = cv2.cvtColor(im, cv2.COLOR_GRAY2RGB)

    scale = cfg.TEST.VISUALIZE.SHOW_SCALE
    im = cv2.resize(im, None, None, fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)

    pred = result.bbox.cpu().numpy()
    labels = result.get_field('labels').cpu().numpy()
    scores = result.get_field('scores').cpu().numpy()
    if cfg.MODEL.TAG_ON:
        tag_scores = result.get_field('tag_scores').cpu().numpy()
        tag_predictions = result.get_field('tag_predictions').cpu().numpy()
    else:
        tag_scores = None
        tag_predictions = None

    # NOTE: contours, recists and diameters not used in new pipeline for Breast CT scans (with mask GT provided)
    contours = None
    recists = None
    diameters = None

    pred *= scale
    overlay, msgs = draw_results(im, pred, labels, scores, tag_predictions=tag_predictions, tag_scores=tag_scores, 
                                 contours=contours, recists=recists, diameters=diameters)
    overlay = print_msg_on_img(overlay, msgs)
    return overlay, msgs


def print_msg_on_img(overlay, msgs):
    txt_height = 20
    msg_im = np.zeros((txt_height*cfg.TEST.VISUALIZE.DETECTIONS_PER_IMG+10, overlay.shape[1], 3), dtype=np.uint8)
    for p in range(len(msgs)):
        msg = msgs[p].split(' | ')
        if cfg.MODEL.TAG_ON: # detection + segmentation + tagging
            msg = msg[0][7:10] + msg[1][:-2] + ': ' + msg[2]
        # elif cfg.MODEL.MASK_ON: # detection + segmentation
        #     msg = msg[0][7:10] + msg[1][:-2]
        else: # just detection
            msg = msg[0][7:10]
        cv2.putText(msg_im, msg, (0, txt_height*(p+1)),
                    cv2.FONT_HERSHEY_DUPLEX, fontScale=.5,
                    color=(255,255,255), thickness=1)
    return np.vstack((overlay, msg_im))